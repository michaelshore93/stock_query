from datetime import datetime
from sqlalchemy import Column, Date, Numeric, create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker
import requests
import key
from operator import itemgetter
import sys
import sqlite3
from openpyxl import Workbook
from openpyxl.chart import Reference, LineChart, BarChart
import pandas as pd

now = datetime.now()
today = "{}-{}-{}".format(now.year, now.month, now.day)

print("""I only have a free API (not paying $35/mo for premium, so \
you can only use my key for certain stocks. \
Apple (AAPL), Microsoft (MSFT), and Goldman Sachs (GS) are some \
that work.\n\nIf you want to check, type \
https://www.quandl.com/api/v3/datasets/EOD/{STOCK TICKER HERE}.json?api_key=AjdDyUMTS3znFrrNsCJh \
into your browser.""")

stock = input("Please enter stock ticker. ").upper()
print(stock)

engine = create_engine("sqlite:///stocks.db")
Base = declarative_base()
conn = sqlite3.connect("stocks.db")
c = conn.cursor()


class Last_Ran_Check:
    def last_row_pr_key():
        """Check for last row of data's primary key."""
        try:
            query = "SELECT * FROM {} ORDER BY Date_yyyy_mm_dd DESC LIMIT 1".format(stock)
            last_date = list(c.execute(query))[0][0]
            return last_date
        except sqlite3.OperationalError:
            last_date = "foobar"
            return last_date
        except IndexError:
            last_date = "foobar"
            return last_date


class JSON_Connection:
    def connect_json():
        """Query JSON data."""
        last_date = Last_Ran_Check.last_row_pr_key()
        url = "https://www.quandl.com/api/v3/datasets/EOD/{}.json?api_key={}".format(stock, key.api_key)
        data_json = requests.get(url).json()
        stock_data = []

        try:
            for element in data_json["dataset"]["data"]:
                    if element[0] == last_date:
                        break
                    dict_set = {}
                    dict_set["Date_yyyy_mm_dd"] = datetime(int(element[0][0:4]), int(element[0][5:7]),int(element[0][8:10]))
                    dict_set["OpenPrice"] = element[8]
                    dict_set["High"] = element[9]
                    dict_set["Low"] = element[10]
                    dict_set["ClosePrice"] = element[11]
                    dict_set["Volume"] = element[12]
                    stock_data.append(dict_set)

            stock_data = sorted(stock_data, key=itemgetter("Date_yyyy_mm_dd"))
        except KeyError:
            table_check = "Invalid ticker."
            print(table_check)
            sys.exit(0)

        return stock_data


class SL_Table(Base):
    """Create the table for the database."""
    table_check = JSON_Connection.connect_json()
    if table_check != "Invalid ticker.":
        __tablename__ = stock
        Date_yyyy_mm_dd = Column(Date, primary_key=True)
        OpenPrice = Column(Numeric)
        High = Column(Numeric)
        Low = Column(Numeric)
        ClosePrice = Column(Numeric)
        Volume = Column(Numeric)


class Table_Bind:
    SL_Table.__table__.create(bind=engine, checkfirst=True)


class DB_Session:
    def json_to_db():
        stock_data = JSON_Connection.connect_json()
        Session = sessionmaker(bind=engine)
        session = Session()

        for element in stock_data:
                row = SL_Table(**element)
                session.add(row)
        session.commit()


class DB_Connection:
    def connect_db():
        """Establish connection with table in SQLite database."""
        mysel = c.execute("SELECT * FROM {}".format(stock))
        connection = sqlite3.connect("stocks.db")
        return mysel, connection


class Workbook:
    wbook_name = "{}_{}.xlsx".format(stock, today)
    wb = Workbook()

    def closeprice():
        """Create an Excel Sheet in our Excel Workbook."""
        ws = Workbook.wb.active
        ws.title = "Close Price"
        mysel, connection = DB_Connection.connect_db()

        """Select relevant data from database and writes it to Excel."""
        cursor = connection.execute("SELECT * FROM {}".format(stock))
        names = list(map(lambda x: x[0], cursor.description))
        idx = 1
        for i in range(6):
            if i == 0 or i == 4:
                ws.cell(row=1, column=idx).value = names[i]
                idx += 1

        counter = 1
        for i, row in enumerate(mysel):
            post = row[0], row[1]
            counter += 1
            ws.append(post)

        """Create line chart of historical data close prices."""
        data = Reference(ws, min_col=2, min_row=counter-1499, max_row=counter, max_col=2)
        cats = Reference(ws, min_col=1, min_row=counter-1498, max_row=counter)

        line = LineChart()
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.title = "Closing Price of Last 1500 Closes"
        line.y_axis.title = "Price"
        line.x_axis.title = "Date"
        line.height = 10
        line.width = 35
        line.legend = None
        ws.add_chart(line, "D10")
        Workbook.wb.save(Workbook.wbook_name)

    def open_close_delta():
        """Create a sheet for Open/Close delta."""
        ws2 = Workbook.wb.create_sheet("Open Close Delta")
        mysel, connection = DB_Connection.connect_db()
        cursor = connection.execute("SELECT * FROM {}".format(stock))
        names = list(map(lambda x: x[0], cursor.description))

        idx = 1
        for i in range(6):
            if i == 0 or i == 1 or i == 4:
                ws2.cell(row=1, column=idx).value = names[i]
                idx += 1

        counter = 1
        "Write relevant data to Excel."
        for i, row in enumerate(mysel):
            post = row[0], row[1], row[4]
            counter += 1
            ws2.append(post)
        Workbook.wb.save(Workbook.wbook_name)

        """Use pandas to create a new field, Open/Close Delta."""
        ws2["D1"] = "Delta"
        df = pd.read_excel(Workbook.wbook_name, "Open Close Delta")
        df["Delta"] = df["OpenPrice"] - df["ClosePrice"]

        for dl in range(2, counter+1):
                ws2["D{}".format(dl)] = float(df["Delta"][dl-2])

        """Plot delta as bar chart."""
        ws2.sheet_view.zoomScale = 85
        bc_three_d = BarChart()
        bc_three_d.type = "col"
        bc_three_d .style = 10
        bc_three_d .title = "Open/Close Delta for Past 30 Trading Days"
        bc_three_d.y_axis.title = "Price"
        bc_three_d.x_axis.title = "Date"
        data = Reference(ws2, min_col=4, min_row=counter-29, max_row=counter, max_col=4)
        cats = Reference(ws2, min_col=1, min_row=counter-28, max_row=counter)
        bc_three_d.add_data(data, titles_from_data=True)
        bc_three_d.set_categories(cats)
        bc_three_d.height = 20
        bc_three_d.width = 30
        bc_three_d.legend = None
        ws2.add_chart(bc_three_d, "F2")
        Workbook.wb.save(Workbook.wbook_name)


c.close()
